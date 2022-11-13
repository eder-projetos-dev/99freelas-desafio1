from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment

def criar_planilha():
    wb = Workbook()
    ws = wb.active

    ws.title = "Página1"
    ws.row_dimensions[1].height = 20
    ws.cell(row = 1, column = 1).font = Font(size = 12, bold = True)
    ws['A1'].alignment = Alignment(horizontal='center')    
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A3' 

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50    
    wb.save('processos.xlsx')    


"""
Recebe: Dados, coluna e linha.
Objetivo: Salvar os dados em uma célula específicada na planilha "Página1"
do arquivo processos.xlsx
Chamada: salvar_xlsx("teste", "A", "1")    
"""

def salvar_xlsx(dados, coluna, linha):
    planilha = load_workbook('processos.xlsx')
    pagina = planilha['Página1']
    pagina[coluna+str(linha)].value=dados; 
    planilha.save('processos.xlsx')
    

"""
Recebe: Nome do arquivo texto.
Objetivo: Ler um arquivo texto e retornar uma lista com as suas linhas.
Chamada: ler_arquivo("arquivo.txt")
"""

def ler_arquivo(arquivo):
        lista=[] 
        arq = open(arquivo, "r")
        texto = arq.readlines()
        for linha in texto :
                lista.append(linha.strip("\n"))
        arq.close()                
        return lista
